import datetime
from io import BytesIO
from unittest.mock import patch

from django.urls import reverse
from dotwiz import DotWiz
from openpyxl import load_workbook

from config.test_bases import UserTestBase

mock_feedback = DotWiz(
    {
        "verbose_rating_name": "Satisfied",
        "verbose_what_didnt_go_so_well": [
            "Process is not clear",
            "Not enough guidance",
            "I was asked for information I don’t have",
            "Other issue",
        ],
        "created_at": datetime.datetime(2022, 11, 24, 14, 19, 4),
        "last_modified": datetime.datetime(2022, 11, 24, 14, 19, 4),
        "logged_in": True,
        "rating": 4,
        "what_didnt_work_so_well": [
            "process_not_clear",
            "not_enough_guidance",
            "asked_for_info_didnt_have",
            "other_issue",
        ],
        "what_didnt_work_so_well_other": "crap",
        "how_could_we_improve_service": "really bad",
        "url": "/dashboard/",
        "journey": "Generic Header",
    }
)


def get_feedbacks():
    return [
        mock_feedback,
    ]


class TestFeedback(UserTestBase):
    @patch("v2_api_client.library.generic.FeedbackAPIClient")
    def setUp(self, mocked_feedback_client) -> None:
        mocked_feedback_client.return_value = get_feedbacks
        self.client.force_login(self.test_user)
        self.response = self.client.get(reverse("export_feedback_objects"))
        self.feedback_export = load_workbook(filename=BytesIO(self.response.content)).active

    def test_headers(self):
        assert self.feedback_export["A1"].value == "Date submitted"
        assert self.feedback_export["C1"].value == "URL"

    def test_feedback_included(self):
        assert self.feedback_export["A2"].value == "24 Nov 2022 at 2:19PM"
        assert self.feedback_export["E2"].value == "Satisfied"
        assert self.feedback_export["K2"].value == "crap"

    def test_only_two_rows(self):
        assert self.feedback_export.max_row == 2

    def test_content_type(self):
        assert self.response.headers["Content-Type"] == "application/vnd.ms-excel"

    def test_file_name(self):
        filename = self.response.headers["Content-Disposition"]
        assert "trs_feedback_export" in filename
        assert (
                datetime.datetime.today()
                .strftime("%d %b %Y at %-I:%M%p")
                .replace(" ", "_")
                .replace(":", "-")
                in filename
        )
        assert ".xlsx" in filename
